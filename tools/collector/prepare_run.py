#!/usr/bin/env python3
"""
prepare_run.py v2  —  Inflection Alpha pipeline run stager
===========================================================
Reads a company's folder from the Google Drive sync, classifies every
file by CONTENT (no renaming ever needed), stages it into the repo's
run structure, converts Excel to CSV, writes the manifest, commits and
pushes. Prints the /run-pipeline line to paste at claude.ai/code.

DRIVE STRUCTURE EXPECTED (documents may be irregularly present; the
script stages whatever exists and flags gaps):
  <company folder>/            deep dive: AR, ~3 concalls, ~2 results,
                               rating update (if any), press releases
                               (if any), a CSV/Excel
    SCREENING FOLDER/          CSV/Excel, sometimes AR + rating copies
    PEER ANALYSIS/<peer>/      ~3 concalls + 1 CSV/Excel per peer

USAGE
  py prepare_run.py SMRUTHI --folder Smruthi --company "Smruthi Organics Ltd" --sector "Pharma / CDMO" --dry-run
  (drop --dry-run when the summary looks right; add --cmp / --mcap to
   skip the interactive prompts)

SETUP (once):  pip install pypdf openpyxl
"""

import argparse, datetime, re, shutil, subprocess, sys
from pathlib import Path

# ----------------------------------------------------------------------
# CONFIG — edit these two paths to match your machine
# ----------------------------------------------------------------------
DRIVE_ROOT = Path(r"G:\My Drive\stock analysis")
REPO_ROOT  = Path(r"C:\Users\SUMIT SHARMA\repos\inflection-pipeline")

PEER_DIR_NAMES      = {"peer analysis", "peer_analysis", "peers"}
SCREENING_DIR_NAMES = {"screening folder", "screening", "screener"}

# ----------------------------------------------------------------------
# PDF classification by content
# ----------------------------------------------------------------------
def pdf_text_head(path: Path, pages: int = 3) -> str:
    try:
        from pypdf import PdfReader
        r = PdfReader(str(path))
        out = []
        for p in r.pages[:pages]:
            try:
                out.append(p.extract_text() or "")
            except Exception:
                pass
        return " ".join(out).lower()
    except Exception as e:
        print(f"    ! could not read {path.name}: {e}")
        return ""

def pdf_page_count(path: Path) -> int:
    try:
        from pypdf import PdfReader
        return len(PdfReader(str(path)).pages)
    except Exception:
        return 0

CONCALL_MARKS = ["earnings conference call", "earnings call transcript",
                 "conference call transcript", "concall", "moderator:"]
RATING_MARKS  = ["crisil", "icra", "care ratings", "careedge",
                 "india ratings", "brickwork", "acuite",
                 "rating rationale", "rating action", "instruments rated",
                 "credit rating"]
RESULT_MARKS  = ["unaudited financial results", "audited financial results",
                 "statement of financial results", "regulation 33",
                 "statement of standalone", "statement of consolidated",
                 "limited review report"]
PRESO_MARKS   = ["investor presentation", "earnings presentation",
                 "corporate presentation"]
PRESS_MARKS   = ["press release", "media release", "intimation under",
                 "disclosure under regulation 30", "outcome of board meeting"]
AR_MARKS      = ["annual report", "board's report", "boards' report",
                 "directors' report", "notice of annual general meeting",
                 "corporate governance report"]

# Order matters: most specific first. "kind" strings == folder names.
def classify_pdf(path: Path) -> str:
    head = pdf_text_head(path)
    pages = pdf_page_count(path)
    def has(marks): return any(m in head for m in marks)
    if has(CONCALL_MARKS): return "concalls"
    if has(RATING_MARKS):  return "rating"
    if has(RESULT_MARKS):  return "results"
    if has(PRESO_MARKS):   return "presentation"
    if has(AR_MARKS) or pages >= 80: return "annual-report"
    if has(PRESS_MARKS):   return "other"
    if pages and pages <= 15: return "results"   # small filings default
    return "_unclassified"

# ----------------------------------------------------------------------
# Excel / Google Sheet handling
# ----------------------------------------------------------------------
def xlsx_to_csvs(xlsx: Path, dest_dir: Path, prefix: str) -> list:
    import csv
    from openpyxl import load_workbook
    wb = load_workbook(str(xlsx), data_only=True, read_only=True)
    written = []
    for ws in wb.worksheets:
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", ws.title).strip("_") or "sheet"
        out = dest_dir / f"{prefix}-{safe}.csv"
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    w.writerow(["" if c is None else c for c in row])
        written.append(out)
    return written

def handle_sheet(path: Path, dest_dir: Path, prefix: str, notes: list):
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        try:
            outs = xlsx_to_csvs(path, dest_dir, prefix)
            print(f"    converted {path.name} -> {len(outs)} csv")
        except Exception as e:
            notes.append(f"Excel convert failed for {path.name}: {e}")
            shutil.copy2(path, dest_dir / path.name)
    elif path.suffix.lower() == ".csv":
        shutil.copy2(path, dest_dir / f"{prefix}-{path.name}")
    elif path.suffix.lower() in {".gsheet", ".gdoc"}:
        notes.append(
            f"'{path.name}' is a Google Sheet POINTER, not data. In Drive web: "
            f"open -> File -> Download -> Microsoft Excel, put the .xlsx in the "
            f"same folder, rerun. Permanent fix: Drive settings -> uncheck "
            f"'Convert uploads' so screener .xlsx stays .xlsx.")

# ----------------------------------------------------------------------
def sanitize(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "-", name).strip("-")

def find_company_folder(sub: str) -> Path:
    hits = [d for d in DRIVE_ROOT.iterdir()
            if d.is_dir() and "output" in d.name.lower()
            and sub.lower() in d.name.lower()]
    if not hits:
        sys.exit(f"No folder under {DRIVE_ROOT} matching *output*{sub}*")
    if len(hits) > 1:
        print("Multiple matches:"); [print("  -", h.name) for h in hits]
        sys.exit("Narrow --folder.")
    return hits[0]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("ticker")
    ap.add_argument("--folder", required=True)
    ap.add_argument("--company", required=True)
    ap.add_argument("--sector", required=True,
                    help='exact Section 1B cap-table row, e.g. "Pharma / CDMO"')
    ap.add_argument("--cmp", type=float)
    ap.add_argument("--mcap", type=float)
    ap.add_argument("--dry-run", action="store_true")
    a = ap.parse_args()

    src = find_company_folder(a.folder)
    print(f"Drive folder : {src}")
    today = datetime.date.today().isoformat()
    run = REPO_ROOT / "runs" / f"{a.ticker.lower()}-{today}"
    inp = run / "inputs"
    kinds = ["annual-report", "results", "rating", "concalls",
             "peer-concalls", "screening", "presentation", "other",
             "_unclassified"]
    dirs = {k: inp / k for k in kinds}
    for d in dirs.values(): d.mkdir(parents=True, exist_ok=True)

    notes, counts = [], {k: 0 for k in kinds}
    seen_names = set()   # dedupe across locations (root vs screening copies)

    def place(pdf: Path, kind: str, prefix: str = ""):
        name = (prefix + "-" if prefix else "") + sanitize(pdf.name)
        # duplicate detection by content shape: same size + page count is
        # the same document, whatever it is named or wherever it sits
        key = (kind, pdf.stat().st_size, pdf_page_count(pdf))
        if key in seen_names:
            print(f"    duplicate skipped ({kind}): {pdf.name}")
            return
        seen_names.add(key)
        shutil.copy2(pdf, dirs[kind] / name)
        counts[kind] += 1
        print(f"    {pdf.name}  ->  {kind}/")

    # ---- deep-dive root: classify every loose PDF ----
    print("\nClassifying deep-dive PDFs (reading first pages)...")
    ar_candidates = []
    for pdf in sorted(src.glob("*.pdf")):
        kind = classify_pdf(pdf)
        if kind == "annual-report": ar_candidates.append(pdf)
        else: place(pdf, kind)

    # ---- screening folder (may hold AR/rating copies + the sheet) ----
    scr = next((d for d in src.iterdir() if d.is_dir()
                and d.name.lower() in SCREENING_DIR_NAMES), None)
    if scr:
        print("Screening folder...")
        for f in sorted(scr.iterdir()):
            if not f.is_file(): continue
            if f.suffix.lower() == ".pdf":
                k = classify_pdf(f)
                if k == "annual-report": ar_candidates.append(f)
                else: place(f, k)
            else:
                handle_sheet(f, dirs["screening"], "screener", notes)
    for f in sorted(src.iterdir()):          # loose sheets in root
        if f.is_file() and f.suffix.lower() != ".pdf":
            handle_sheet(f, dirs["screening"], "financials", notes)

    # ---- annual report: keep the most complete candidate ----
    if ar_candidates:
        best = max(ar_candidates, key=pdf_page_count)
        place(best, "annual-report")
        for other in {c for c in ar_candidates if c.resolve() != best.resolve()}:
            notes.append(f"AR candidate '{other.name}' skipped "
                         f"({pdf_page_count(other)}p vs kept {pdf_page_count(best)}p). "
                         f"Verify the kept file is the FULL annual report.")

    # ---- peer analysis: one subfolder per peer ----
    peers = next((d for d in src.iterdir() if d.is_dir()
                  and d.name.lower() in PEER_DIR_NAMES), None)
    if peers:
        print("Peer analysis folders...")
        for sub in sorted(p for p in peers.iterdir() if p.is_dir()):
            is_main = (a.company.lower() in sub.name.lower()
                       or a.ticker.lower() in sub.name.lower())
            tag = sanitize(sub.name)
            for f in sorted(sub.rglob("*")):
                if not f.is_file(): continue
                if f.suffix.lower() == ".pdf":
                    if is_main:
                        # main co's own folder: concalls belong to main set;
                        # dedupe guard handles overlap with deep-dive root
                        place(f, classify_pdf(f))
                    else:
                        place(f, "peer-concalls", tag)
                else:
                    handle_sheet(f, dirs["screening"],
                                 ("main" if is_main else tag), notes)

    # ---- contract check (rating now OPTIONAL; gaps flagged not fatal) ----
    concalls_available = counts["concalls"] >= 3
    problems, warnings = [], []
    total_docs = sum(counts[k] for k in kinds if k != "screening")
    if total_docs == 0 and counts["screening"] == 0:
        problems.append("no documents found at all: nothing to analyse")
    if counts["annual-report"] == 0:
        warnings.append("no annual report: notes triple-pass and AR deep dive "
                        "will be skipped, gap recorded; verdict likely "
                        "INSUFFICIENT EVIDENCE unless other sources are strong")
    elif counts["annual-report"] > 1:
        warnings.append(f"{counts['annual-report']} ARs staged; pipeline uses "
                        f"the most recent/most complete")
    if counts["results"] == 0:
        warnings.append("no results PDFs: Gate 0 relies on screening data "
                        "only; latest-quarter fields unresolved")
    elif counts["results"] == 1:
        warnings.append("only 1 results PDF: trend fields limited, gap recorded")
    elif counts["results"] > 3:
        warnings.append(f"results has {counts['results']}; pipeline uses the "
                        f"3 most recent, extras are fine")
    if counts["rating"] == 0:
        warnings.append("no rating PDF found: run proceeds, gap recorded; "
                        "Pillar 2 structural/growth determination loses its "
                        "strongest evidence source")
    elif counts["rating"] > 1:
        warnings.append(f"{counts['rating']} rating PDFs; pipeline uses most recent")
    if 0 < counts["concalls"] < 3:
        warnings.append(f"only {counts['concalls']} concall(s); manifest sets "
                        f"concalls_available: false, no-concall mode applies")
        concalls_available = False
    for k in ("other", "_unclassified"):
        if counts[k] == 0:
            shutil.rmtree(dirs[k], ignore_errors=True)
        elif k == "_unclassified":
            warnings.append(f"{counts[k]} file(s) in _unclassified/: inspect and "
                            f"move manually before pushing, or leave (pipeline ignores)")

    # ---- manifest ----
    cmp_v  = a.cmp  if a.cmp  is not None else float(input("CMP (₹): ").strip() or 0)
    mcap_v = a.mcap if a.mcap is not None else float(input("Market cap (₹ Cr): ").strip() or 0)
    (run / "manifest.yaml").write_text(
f"""company: {a.company}
ticker: {a.ticker.upper()}
cmp: {cmp_v}
market_cap_cr: {mcap_v}
run_date: {today}
run_type: full
concalls_available: {str(concalls_available).lower()}
sector_cap_row: "{a.sector}"
notes: "Staged by prepare_run.py v2 from Drive folder '{src.name}'."
""", encoding="utf-8")

    # ---- summary + confirm ----
    print("\n================ STAGING SUMMARY ================")
    for k, v in counts.items():
        if v: print(f"  {k:15s} {v}")
    print(f"  concalls_available: {concalls_available}")
    for w in warnings: print(f"  ⚠ {w}")
    for n in notes:    print(f"  ⚠ {n}")
    for p in problems: print(f"  ✗ BLOCKER: {p}")
    print(f"  run folder: {run}")
    print("=================================================")
    if problems:
        print("\nFix the blockers (add/remove files in Drive or directly in the")
        print("run folder), then rerun. Nothing pushed.")
        return
    if a.dry_run:
        print("\n--dry-run: staged locally, nothing committed."); return
    if input("\nPush to GitHub? [y/N] ").strip().lower() != "y":
        print("Staged locally, not pushed."); return

    # ---- git ----
    def git(*args):
        r = subprocess.run(["git", *args], cwd=REPO_ROOT,
                           capture_output=True, text=True)
        if r.returncode != 0:
            sys.exit(f"git {' '.join(args)} failed:\n{r.stderr}")
        return r.stdout
    git("pull", "--rebase", "origin", "main")
    git("add", str(run.relative_to(REPO_ROOT)))
    git("commit", "-m", f"run inputs: {a.ticker.upper()} {today}")
    git("push", "origin", "main")

    print("\n✅ Pushed. Now start a session at claude.ai/code and paste:\n")
    print(f"   /run-pipeline runs/{run.name}\n")

if __name__ == "__main__":
    main()
