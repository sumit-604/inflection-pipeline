#!/usr/bin/env python3
"""
collect_to_repo.py v3  —  companies.txt -> screener.in -> GitHub. Done.
========================================================================
Works exactly like collect_batch.py: reads companies.txt from this
folder (the block Claude gives you, pasted into notepad, saved). Then:

    py collect_to_repo.py

That's the whole interface. No flags, no prompts.
- CMP, market cap, company name: scraped from the screener page itself.
- Sector cap row: picked automatically by a keyword map (printed in the
  summary so you can see what it chose; edit manifest.yaml if wrong).
- Results/rating PDFs: not on screener. Push happens without them; add
  them later directly on GitHub (drag into runs/<folder>/inputs/results/
  and /rating/) or drop them in the local run folder and run
  `py collect_to_repo.py --push-again`.

companies.txt format (same as always):
    APEXECO:https://www.screener.in/company/APEXECO/
    PEER:APEXECO:https://www.screener.in/company/CEWATER/consolidated/
    PEER:APEXECO:https://www.screener.in/company/FELIX/consolidated/

Options (rarely needed):
    --dry-run       stage + classify, don't push
    --push-again    re-stage/commit the latest run folder (after adding
                    manual PDFs locally) and push
SETUP: same folder as screener_collect.py. pip install pypdf openpyxl
"""

import argparse, csv, datetime, re, shutil, subprocess, sys
from pathlib import Path

# Windows: when stdout is redirected (pipe/file) Python falls back to the
# locale codec (cp1252) which can't encode ₹ and other non-latin-1 chars,
# crashing on the summary print. Force UTF-8 so redirected runs don't die.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

SCRIPT_DIR = Path(__file__).parent
COLLECTOR  = SCRIPT_DIR / "screener_collect.py"
COMPANIES  = SCRIPT_DIR / "companies.txt"
REPO_ROOT  = Path(r"C:\Users\SUMIT SHARMA\repos\inflection-pipeline")

# sector cap row picked by first keyword hit on screener's page text /
# company name; edit freely. Fallback prints a reminder.
SECTOR_MAP = [
    (["pharma", "drug", "api", "laborator", "lifescience", "healthcare",
      "formulation"], "Pharma / CDMO"),
    (["hospital", "diagnostic", "dialysis"], "Hospitals / dialysis / healthcare services"),
    (["chemical", "organics", "specialit", "specialty", "agrochem"], "Specialty chemicals"),
    (["cable", "wire", "conductor", "transformer", "switchgear", "electrical",
      "heat exchanger", "industrial"], "Cables / Industrial products"),
    (["software", "technolog", "infotech", "it services", "digital", "data analytics"],
     "Platform / SaaS / IT services"),
    (["bank", "finance", "nbfc", "credit", "capital", "microfinance"],
     "Banks / NBFCs / MFIs"),
    (["infra", "construction", "engineering procurement", "epc", "projects"],
     "EPC / Civil construction"),
    (["logistic", "cargo", "shipping", "freight"], "Logistics (asset-light)"),
    (["hotel", "resort", "hospitality"], "Hotels"),
    (["textile", "apparel", "garment", "fashion"], "Branded apparel / FMCG"),
    (["food", "agro", "dairy", "edible", "sugar", "rice"], "Agri processing"),
    (["cement", "tiles", "ceramic", "pipes", "sanitary"], "Building materials"),
    (["packag", "container", "laminates"], "Packaging"),
    (["real estate", "realty", "developer", "properties"], "Real estate"),
    (["defence", "defense", "aerospace"], "Defence / strategic"),
    (["telecom", "network equipment"], "Telecom equipment"),
    (["recycl", "metal", "steel", "alloy", "casting", "forging"],
     "Recycling / Manufacturing"),
]

# ---------------- pdf classification ----------------
def pdf_text_head(p, pages=3):
    try:
        from pypdf import PdfReader
        r = PdfReader(str(p))
        return " ".join((pg.extract_text() or "") for pg in r.pages[:pages]).lower()
    except Exception:
        return ""

def pdf_pages(p):
    try:
        from pypdf import PdfReader
        return len(PdfReader(str(p)).pages)
    except Exception:
        return 0

MARKS = {
    "concalls": ["earnings conference call", "earnings call transcript",
                 "conference call transcript", "concall", "moderator:"],
    "rating":   ["crisil", "icra", "care ratings", "careedge", "india ratings",
                 "brickwork", "acuite", "rating rationale", "rating action",
                 "instruments rated", "credit rating"],
    "results":  ["unaudited financial results", "audited financial results",
                 "statement of financial results", "regulation 33",
                 "statement of standalone", "statement of consolidated",
                 "limited review report"],
    "presentation": ["investor presentation", "earnings presentation",
                     "corporate presentation"],
    "other":    ["press release", "media release", "intimation under",
                 "disclosure under regulation 30", "outcome of board meeting"],
}
AR_MARKS = ["annual report", "board's report", "boards' report",
            "directors' report", "notice of annual general meeting",
            "corporate governance report"]

def classify(p):
    n = p.name.lower()
    if "concall" in n or "transcript" in n: return "concalls"
    if "annual_report" in n or n == "ar.pdf": return "annual-report"
    if "presentation" in n: return "presentation"
    head, pages = pdf_text_head(p), pdf_pages(p)
    for kind, marks in MARKS.items():
        if any(m in head for m in marks): return kind
    if any(m in head for m in AR_MARKS) or pages >= 80: return "annual-report"
    if pages and pages <= 15: return "results"
    return "_unclassified"

def xlsx_to_csvs(x, dest, prefix):
    from openpyxl import load_workbook
    wb = load_workbook(str(x), data_only=True, read_only=True)
    n = 0
    try:
        for ws in wb.worksheets:
            safe = re.sub(r"[^A-Za-z0-9._-]+", "_", ws.title).strip("_") or "sheet"
            with open(dest / f"{prefix}-{safe}.csv", "w", newline="",
                      encoding="utf-8") as f:
                w = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        w.writerow(["" if c is None else c for c in row])
            n += 1
    finally:
        wb.close()   # Windows: unclosed handles make files undeletable
    return n

def sanitize(s): return re.sub(r"[^A-Za-z0-9._-]+", "-", s).strip("-")

# ---------------- companies.txt ----------------
def parse_companies():
    if not COMPANIES.exists():
        sys.exit("companies.txt not found next to this script.")
    main_ticker, main_url, peers = None, None, []
    for line in COMPANIES.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"): continue
        if line.upper().startswith("PEER:"):
            parts = line.split(":", 2)
            if len(parts) == 3: peers.append(parts[2].strip())
        else:
            if ":" in line and not line.lower().startswith("http"):
                t, u = line.split(":", 1)
                main_ticker, main_url = t.strip().upper(), u.strip()
            else:
                main_url = line
                m = re.search(r"/company/([^/]+)/", line)
                main_ticker = m.group(1).upper() if m else "COMPANY"
    if not main_url: sys.exit("No main company line found in companies.txt.")
    return main_ticker, main_url, peers

# ---------------- scrape name / cmp / mcap from screener page ----------------
def scrape_company_facts(url):
    """Return (name, cmp, mcap_cr) from the public screener page."""
    import requests
    name, cmp_v, mcap = None, 0.0, 0.0
    try:
        html = requests.get(url, timeout=30, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}).text
        m = re.search(r"<h1[^>]*>\s*([^<\n]{3,80})", html)
        if m: name = re.sub(r"\s+", " ", m.group(1)).strip()
        flat = re.sub(r"<[^>]+>", " ", html)
        flat = re.sub(r"\s+", " ", flat)
        m = re.search(r"Current Price\s*₹?\s*([\d,]+(?:\.\d+)?)", flat)
        if m: cmp_v = float(m.group(1).replace(",", ""))
        m = re.search(r"Market Cap\s*₹?\s*([\d,]+(?:\.\d+)?)\s*Cr", flat)
        if m: mcap = float(m.group(1).replace(",", ""))
        return name, cmp_v, mcap, flat.lower()
    except Exception as e:
        print(f"  ⚠ could not scrape screener page ({str(e)[:60]}); "
              f"manifest gets zeros, fix by editing manifest.yaml")
        return name, cmp_v, mcap, ""

def pick_sector(name, page_text):
    hay = ((name or "") + " " + page_text[:4000]).lower()
    for keys, row in SECTOR_MAP:
        if any(k in hay for k in keys): return row
    return "Recycling / Manufacturing"   # conservative fallback, flagged

# ---------------- git ----------------
def git(*args):
    r = subprocess.run(["git", *args], cwd=REPO_ROOT,
                       capture_output=True, text=True)
    if r.returncode != 0:
        sys.exit(f"git {' '.join(args)} failed:\n{r.stderr}")

def git_try(*args):
    """Best-effort git command; never aborts the script."""
    subprocess.run(["git", *args], cwd=REPO_ROOT,
                   capture_output=True, text=True)

def push_run(run, ticker, today):
    git_try("rebase", "--abort")          # clear any stuck rebase
    git("add", "-A")                      # stage EVERYTHING incl .gitignore
    git_try("commit", "-m", f"run inputs: {ticker} {today}")  # ok if empty
    git("pull", "--rebase", "origin", "main")
    git("push", "origin", "main")
    print(f"\n✅ Pushed. At claude.ai/code paste:\n\n   /run-pipeline {ticker}\n")
    print("Results/rating PDFs missing? Add them any time: drag into")
    print(f"runs/{run.name}/inputs/results/ (or /rating/) on github.com,")
    print("or drop locally into the run folder and: py collect_to_repo.py --push-again")

# ---------------- main ----------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--push-again", action="store_true")
    a = ap.parse_args()

    if a.push_again:
        runs = sorted((REPO_ROOT / "runs").iterdir(), key=lambda d: d.name)
        runs = [d for d in runs if d.is_dir() and not d.name.startswith("_")]
        if not runs: sys.exit("No run folders found.")
        run = runs[-1]
        t = run.name.split("-")[0].upper()
        print(f"Re-pushing latest run folder: {run.name}")
        push_run(run, t, datetime.date.today().isoformat())
        return

    ticker, url, peer_urls = parse_companies()
    print(f"Main:  {ticker}  {url}")
    for u in peer_urls: print(f"Peer:  {u}")

    print("\nFetching company facts from screener page...")
    name, cmp_v, mcap, page_text = scrape_company_facts(url)
    name = name or ticker.title()
    sector = pick_sector(name, page_text)
    print(f"  name:   {name}\n  cmp:    ₹{cmp_v}\n  mcap:   ₹{mcap} Cr"
          f"\n  sector: {sector}  (edit manifest.yaml if wrong)")

    gi = REPO_ROOT / ".gitignore"
    gi_text = gi.read_text(encoding="utf-8") if gi.exists() else ""
    if "_download/" not in gi_text:
        gi.write_text(gi_text.rstrip() + "\n_download/\n", encoding="utf-8")

    today = datetime.date.today().isoformat()
    run   = REPO_ROOT / "runs" / f"{ticker.lower()}-{today}"
    stage = run / "_download"
    stage.mkdir(parents=True, exist_ok=True)

    print(f"\nDownloading main company -> {stage}")
    subprocess.run([sys.executable, str(COLLECTOR), url,
                    "--output-dir", str(stage)], cwd=SCRIPT_DIR)
    peer_dirs = {}
    for u in peer_urls:
        m = re.search(r"/company/([^/]+)/", u)
        pt = (m.group(1).upper() if m else "PEER")
        pd = stage / "peers" / pt
        pd.mkdir(parents=True, exist_ok=True)
        peer_dirs[pt] = pd
        print(f"\nDownloading peer {pt} -> {pd}")
        subprocess.run([sys.executable, str(COLLECTOR), u,
                        "--output-dir", str(pd)], cwd=SCRIPT_DIR)

    # classify
    print("\nClassifying...")
    kinds = ["annual-report", "results", "rating", "concalls",
             "peer-concalls", "screening", "presentation", "other",
             "_unclassified"]
    inp  = run / "inputs"
    dirs = {k: inp / k for k in kinds}
    for d in dirs.values(): d.mkdir(parents=True, exist_ok=True)
    counts = {k: 0 for k in kinds}
    seen, warnings, ar_cand = set(), [], []

    def place(pdf, kind, prefix=""):
        key = (kind, pdf.stat().st_size, pdf_pages(pdf))
        if key in seen: return
        seen.add(key)
        shutil.copy2(pdf, dirs[kind] /
                     ((prefix + "-" if prefix else "") + sanitize(pdf.name)))
        counts[kind] += 1
        print(f"    {pdf.name} -> {kind}/")

    for f in sorted(stage.iterdir()):
        if f.is_file():
            if f.suffix.lower() == ".pdf":
                k = classify(f)
                ar_cand.append(f) if k == "annual-report" else place(f, k)
            elif f.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
                counts["screening"] += xlsx_to_csvs(f, dirs["screening"], "screener")
    if ar_cand:
        # keep the two most recent years (year taken from the filename), using
        # page count only as a tie-break / fallback when no year is present.
        # "most pages" alone silently dropped a newer, image-heavy AR (FY26,
        # 120pp) in favour of an older, longer one (FY25, 242pp).
        def ar_rank(p):
            m = re.search(r"(20\d\d)", p.name)
            return (int(m.group(1)) if m else 0, pdf_pages(p))
        for ar in sorted(ar_cand, key=ar_rank, reverse=True)[:2]:
            place(ar, "annual-report")
    for pt, pd in peer_dirs.items():
        for f in sorted(pd.iterdir()):
            if not f.is_file(): continue
            if f.suffix.lower() == ".pdf" and classify(f) == "concalls":
                place(f, "peer-concalls", pt)
            elif f.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
                counts["screening"] += xlsx_to_csvs(f, dirs["screening"], pt)

    concalls_available = counts["concalls"] >= 3
    if counts["annual-report"] == 0:
        warnings.append("no AR found on screener: AR stages will be skipped")
    if counts["results"] == 0:
        warnings.append("no results PDFs yet (screener has none): add later on "
                        "github or locally + --push-again")
    if counts["rating"] == 0:
        warnings.append("no rating PDF yet: add later the same way")
    if 0 < counts["concalls"] < 3:
        warnings.append(f"only {counts['concalls']} concalls -> no-concall mode")
        concalls_available = False
    for k in ("other", "_unclassified"):
        if counts[k] == 0: shutil.rmtree(dirs[k], ignore_errors=True)

    (run / "manifest.yaml").write_text(
f"""company: {name}
ticker: {ticker}
cmp: {cmp_v}
market_cap_cr: {mcap}
run_date: {today}
run_type: full
concalls_available: {str(concalls_available).lower()}
sector_cap_row: "{sector}"
notes: "Collected from screener.in by collect_to_repo.py v3. Sector row auto-picked; verify."
""", encoding="utf-8")

    print("\n================ SUMMARY ================")
    for k, v in counts.items():
        if v: print(f"  {k:15s} {v}")
    print(f"  concalls_available: {concalls_available}")
    print(f"  sector_cap_row:     {sector}")
    for w in warnings: print(f"  ⚠ {w}")
    print(f"  run folder: {run}")
    print("=========================================")

    shutil.rmtree(stage, ignore_errors=True)
    if stage.exists():
        print(f"  ⚠ staging not fully deleted ({stage}); harmless, it is "
              f"gitignored, but you may remove it manually later")
    if a.dry_run:
        print("\n--dry-run: staged locally, nothing pushed."); return
    print("\nPushing to GitHub automatically...")
    push_run(run, ticker, today)

if __name__ == "__main__":
    main()
